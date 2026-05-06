import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('Consuntivo aprile 2026.xlsx', data_only=True)
ws = wb.active

headers = [c.value for c in next(ws.iter_rows(max_row=1))]
print('Colonne:', headers)
print()

idx = {h: i for i, h in enumerate(headers)}
col_imp = idx.get('Importo')
col_cli = idx.get('Cliente')
col_data = idx.get('Data')
col_prod = idx.get('Prodotto')

if col_imp is None:
    print('!!! Colonna "Importo" non trovata')
    raise SystemExit

tot = 0.0
n = 0
per_cli = defaultdict(float)
per_prod = defaultdict(float)
per_mese = defaultdict(float)
righe_zero = 0
righe_strane = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[col_cli] is None and row[col_imp] is None:
        continue
    imp = row[col_imp]
    if imp is None:
        righe_zero += 1
        continue
    try:
        imp_f = float(imp)
    except Exception:
        righe_strane.append(row)
        continue
    tot += imp_f
    n += 1
    cli = (row[col_cli] or '').strip() if isinstance(row[col_cli], str) else str(row[col_cli] or '')
    per_cli[cli] += imp_f
    if col_prod is not None:
        per_prod[row[col_prod] or 'NULL'] += imp_f
    if col_data is not None and row[col_data] is not None:
        d = row[col_data]
        if hasattr(d, 'month'):
            per_mese[f"{d.year}-{d.month:02d}"] += imp_f

print(f'Righe valide: {n}')
print(f'Righe con importo vuoto: {righe_zero}')
print(f'TOTALE Importo: {tot:.2f} EUR')
print()

print('--- Per mese (data movimento) ---')
for k in sorted(per_mese):
    print(f'  {k}: {per_mese[k]:.2f}')
print()

print('--- Top 15 clienti per importo ---')
for c, v in sorted(per_cli.items(), key=lambda x: -x[1])[:15]:
    print(f'  {v:>10.2f}  {c}')
print()

print('--- Per prodotto ---')
for p, v in sorted(per_prod.items(), key=lambda x: -x[1]):
    print(f'  {v:>10.2f}  {p}')
